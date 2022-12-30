import sqlparse
from mo_sql_parsing import parse_sqlserver as parse
from openpyxl import Workbook
import openpyxl
from obj.record import Record
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook
from schemaGraph import SchemaGraph
import os

def getDbNameList():
    db_list = os.listdir('spider/database/')
    return db_list

def parseSQL(sql):
    parsed = parse(sql)
    evaluateError(parsed)

# evaluate the case that tables in SELECT and WHERE clause does not appear in FROM

def extractTableFromSelect(item):
    if isinstance(item['value'], str):
        return item['value'].split('.')[0]
        # check_list.append(item['value'].split('.')[0])
    else:
        key = list(item['value'].keys())[0]
        content = item['value'][key]
        if content != "*":
            return content.split('.')[0]
            # check_list.append(content.split('.')[0])

def extractTableFromFROM(from_clause, join_list):
    if type(from_clause) == str:
        join_list.append(from_clause)
        return
    for item in from_clause:
        if type(item) == str:
            join_list.append(item)
        else:
            join_list.append(item["join"])

def extractTableFromWhere(item, check_list):
    # print(item)
    key = list(item.keys())[0]
    if type(item[key]) == list and type(item[key][0]) == dict:
        for i in item[key]:
            key_i = list(i.keys())[0]
            table = i[key_i][0].split('.')[0]
            check_list.append(table)
    else:
        # process directly
        table = item[key][0].split('.')[0]
        check_list.append(table)

# def isSublist(a,b):
#     for i in range(len(b)-len(a)+1):
#         if b[i:i+len(a)] == a:
#             return True
#     return False

def checkExistance(check_list, join_list):
    isExist = True
    for item in check_list:
        if item != None and item not in join_list:
            isExist = False
    return isExist

# def checkPath(join_list, graph_dict):
#     path_forward = []
#     path_backward = []
#     stack = []
#
#     root = join_list[0]
#
#     # forward path
#     stack.append(root)
#     while len(stack) != 0:
#         table = stack.pop()
#         path_forward.append(table)
#         ref_tables = graph_dict[table]['to']
#         for item in ref_tables:
#             stack.append(item[0])
#     # print('forward: ', path_forward)
#
#     # backward path
#     stack.append(root)
#     while len(stack) != 0:
#         table = stack.pop()
#         path_backward.insert(0, table)
#         ref_by_tables = graph_dict[table]['from']
#         for item in ref_by_tables:
#             stack.append(item)
#     # print('backward: ', path_backward)
#
#     path =  path_backward[:-1] + path_forward
#     # print('full_path: ', path)
#
#     if not isSublist(join_list, path):
#         return False
#     return True

def checkPath(join_list, graph_dict):
    for i, table in enumerate(join_list):
        # already the last item, return true
        if i + 1 == len(join_list):
            return True
        # check the correctness of single step
        candidate_next = graph_dict[table]['to'] + graph_dict[table]['from']
        if join_list[i+1] not in candidate_next:
            return False

def evaluateError(result, graph_dict):
    check_list = []
    join_list = []
    #print(result)
    select = result['select']
    if type(select) == list:
        for item in select:
            check_list.append(extractTableFromSelect(item))
    else:
        check_list.append(extractTableFromSelect(select))
    if 'where' in result:
        # print(result['where'])
        extractTableFromWhere(result['where'], check_list)

        # for item in result['where']:
        #     join_list.append(item['value'].split('.')[0])
    extractTableFromFROM(result['from'], join_list)

    # print("check list: ", check_list)
    # print("join list: ", join_list)
    isPath = checkPath(join_list, graph_dict)
    isExist = checkExistance(check_list, join_list)
    return (isExist, isPath)

def readFile():
    wb = load_workbook('error_predictions_bert.xlsx')
    db_list = getDbNameList()

    count_total = 0
    count_false_existance = 0
    count_false_path = 0
    count_both = 0

    for name in db_list:
        try:
            wb_sheet = wb[name]
            graph = SchemaGraph(name)
            graph_dict = graph.getGraph()
        except:
            continue
        for i in range(2, wb_sheet.max_row + 1):
            q1 = wb_sheet.cell(row=i, column=column_index_from_string('D')).value
            try:
                # parseSQL(q1)
                result = evaluateError(parse(q1), graph_dict)

                if result[0] == False:
                    count_false_existance += 1
                if result[1] == False:
                    count_false_path += 1
                if result[0] == False and result[1] == False:
                    count_both += 1
            except:
                continue
                #print(q1)
            count_total += 1
    print('Error of existance: ', count_false_existance, '/', count_total)
    print('Error of join path: ', count_false_path, '/', count_total)
    print('Error of both: ', count_both, '/', count_total)

readFile()

# SELECT countrylanguage.CountryCode FROM countrylanguage WHERE countrylanguage.Language = "value"
# SELECT country.Name FROM country WHERE country.Continent = \"value\" AND country.Population > \"value\"
# SELECT COUNT(*) FROM car_names JOIN cars_data WHERE cars_data.Year = \"value\"
# parseSQL("SELECT COUNT(*) FROM car_names WHERE cars_data.Year = \"value\"")




# l1 = ["c", "b",  "c", "d"]
# l2 = ["c", "a", "c", "b", "c", "c", "d", "d", "f", "e"]
#
# print(isSublist(l1, l2))