from openpyxl import Workbook
import openpyxl
from obj.record import Record
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook

version = "bert"
input_path = f"output_{version}.txt".format(version=version)
output_path = f"error_predictions_{version}.xlsx".format(version=version)
def readFile():
    with open(input_path, 'r') as f:
        lines = f.readlines()
        dict = {} # {db_name: [record1, record2, ...]}
        info = []
        for line in lines:
            if line.strip() == "":
                #save info
                saveInfo(info, dict)
                info.clear()
            else:
                seg = line.split(':')
                info.append(seg[1])
        return dict

def saveInfo(info, dict):
    db = info[4].strip()
    isTrue = True if info[5].strip() == 'True' else False
    record = Record(info[0].strip(), info[1].strip(), info[2].strip(), info[3].strip(), info[4].strip(), isTrue)
    if db in dict:
        dict[db].append(record)
    else:
        recordList = []
        recordList.append(record)
        dict.update({db: recordList})


def outputError(dict):
    wb = Workbook()
    for key, value in dict.items():
        sheet = wb.create_sheet(key)
        sheet.title = key

        sheet.cell(row=1, column=column_index_from_string('A')).value = 'Level'
        sheet.cell(row=1, column=column_index_from_string('B')).value = 'Question'
        sheet.cell(row=1, column=column_index_from_string('C')).value = 'Gold'
        sheet.cell(row=1, column=column_index_from_string('D')).value = 'Predict'
        row = 2
        for record in value:

            if record.correct == False:
                sheet.cell(row=row, column=column_index_from_string('A')).value = record.hardness
                sheet.cell(row=row, column=column_index_from_string('B')).value = record.question
                sheet.cell(row=row, column=column_index_from_string('C')).value = record.gold
                sheet.cell(row=row, column=column_index_from_string('D')).value = record.predict
                sheet.cell(row=row, column=column_index_from_string('E')).value = record.correct
                row = row + 1

    wb.save(filename=output_path)

def outputJoin(dict):
    wb = load_workbook(output_path)
    sheet = wb.create_sheet("statistics_join", 0)

    sheet.cell(row=1, column=column_index_from_string('A')).value = 'DB'
    sheet.cell(row=1, column=column_index_from_string('B')).value = '#Join'
    sheet.cell(row=1, column=column_index_from_string('C')).value = '#Eazy_T'
    sheet.cell(row=1, column=column_index_from_string('D')).value = '#Eazy_F'
    sheet.cell(row=1, column=column_index_from_string('E')).value = '#Medium_T'
    sheet.cell(row=1, column=column_index_from_string('F')).value = '#Medium_F'
    sheet.cell(row=1, column=column_index_from_string('G')).value = '#Hard_T'
    sheet.cell(row=1, column=column_index_from_string('H')).value = '#Hard_F'
    sheet.cell(row=1, column=column_index_from_string('I')).value = '#Extra_T'
    sheet.cell(row=1, column=column_index_from_string('J')).value = '#Extra_F'

    row = 2
    print(dict)
    for key, value in dict.items():
        max_join = 0
        for record in value:
            join_count = record.gold.upper().count("JOIN")
            if join_count > max_join:
                max_join = join_count
        list = [None] * max_join
        #list index: join count
        #list item: {eazy: [t, f], medium: [t, f]}
        for record in value:
            count = record.gold.upper().count("JOIN")
            if count == 0:
                continue
            index = 0
            if record.correct == False:
                index = 1
            if list[count-1] != None:
                list[count-1][record.hardness][index] += 1
            else:
                # init dictionary
                init_dict = {"easy": [0, 0], "medium": [0, 0], "hard": [0, 0], "extra": [0, 0]}
                init_dict[record.hardness][index] += 1
                list[count - 1] = init_dict
        print(list)
        for i in range(len(list)):
            sheet.cell(row=row, column=column_index_from_string('A')).value = key
            sheet.cell(row=row, column=column_index_from_string('B')).value = i+1

            if list[i] != None:
                sheet.cell(row=row, column=column_index_from_string('C')).value = list[i]["easy"][0]
                sheet.cell(row=row, column=column_index_from_string('D')).value = list[i]["easy"][1]
                sheet.cell(row=row, column=column_index_from_string('E')).value = list[i]["medium"][0]
                sheet.cell(row=row, column=column_index_from_string('F')).value = list[i]["medium"][1]
                sheet.cell(row=row, column=column_index_from_string('G')).value = list[i]["hard"][0]
                sheet.cell(row=row, column=column_index_from_string('H')).value = list[i]["hard"][1]
                sheet.cell(row=row, column=column_index_from_string('I')).value = list[i]["extra"][0]
                sheet.cell(row=row, column=column_index_from_string('J')).value = list[i]["extra"][1]

            row += 1

    wb.save(filename=output_path)

def saveToFile(results):
    wb = load_workbook(output_path)
    sheet = wb['Sheet']

    sheet.cell(row=1, column=column_index_from_string('A')).value = 'db'
    sheet.cell(row=1, column=column_index_from_string('B')).value = 'C_true'
    sheet.cell(row=1, column=column_index_from_string('C')).value = 'C_false'
    sheet.cell(row=1, column=column_index_from_string('D')).value = 'error_rate'

    row = 2
    for result in results:
        sheet.cell(row=row, column=column_index_from_string('A')).value = result[0]
        sheet.cell(row=row, column=column_index_from_string('B')).value = result[1]
        sheet.cell(row=row, column=column_index_from_string('C')).value = result[2]
        sheet.cell(row=row, column=column_index_from_string('D')).value = result[3]
        row = row + 1
    wb.save(filename=output_path)

def getStatistics(dict):
    results = [] #[[db, count_true, count_false, error_rate], ...]
    for key, value in dict.items():
        count_t = 0
        count_f = 0
        for record in value:
            if record.correct:
                count_t = count_t + 1
            else:
                count_f = count_f + 1
        error_rate = count_f / (count_t + count_f)
        list = [key, count_t, count_f, error_rate]
        results.append(list)
    # sort [1] count_true, [2] count false, [3] error rate
    results = sorted(results, key=lambda x: x[3], reverse=True)
    # print(results)
    saveToFile(results)

def saveExtraInfo(results):
    wb = load_workbook(output_path)
    sheet = wb['Sheet']

    #row = 2
    for row in range(2, sheet.max_row + 1):
        db = sheet.cell(row=row, column=column_index_from_string('A')).value
        print(results[db])
        sheet.cell(row=row, column=column_index_from_string('E')).value = results[db]['easy'][0]
        sheet.cell(row=row, column=column_index_from_string('F')).value = results[db]['easy'][1]

        sheet.cell(row=row, column=column_index_from_string('G')).value = results[db]['medium'][0]
        sheet.cell(row=row, column=column_index_from_string('H')).value = results[db]['medium'][1]

        if 'hard' in results[db]:
            sheet.cell(row=row, column=column_index_from_string('I')).value = results[db]['hard'][0]
            sheet.cell(row=row, column=column_index_from_string('J')).value = results[db]['hard'][1]
        else:
            sheet.cell(row=row, column=column_index_from_string('I')).value = 0
            sheet.cell(row=row, column=column_index_from_string('J')).value = 0

        if 'extra' in results[db]:
            sheet.cell(row=row, column=column_index_from_string('K')).value = results[db]['extra'][0]
            sheet.cell(row=row, column=column_index_from_string('L')).value = results[db]['extra'][1]
        else:
            sheet.cell(row=row, column=column_index_from_string('K')).value = 0
            sheet.cell(row=row, column=column_index_from_string('L')).value = 0
        row = row + 1
    wb.save(filename=output_path)

def getExtraSta(dict):
    results = {}
    for key, value in dict.items():
        dict = {} #{hardness:[c_true, c_false]}
        for record in value:
            if record.hardness in dict:
                if record.correct == True:
                    dict[record.hardness][0] += 1
                else:
                    dict[record.hardness][1] += 1
            else:
                if record.correct == True:
                    dict.update({record.hardness: [1, 0]})
                else:
                    dict.update({record.hardness: [0, 1]})
            results.update({key: dict})
    # sort [1] count_true, [2] count false, [3] error rate
    # results = sorted(results, key=lambda x: x[3], reverse=True)
    print(results)
    saveExtraInfo(results)

dict = readFile()

# print a excel doc contains all failure case
outputError(dict)

# statistics of error rate of each db, find the one with the largest rate

# getStatistics(dict)
# getExtraSta(dict)
#
# outputJoin(dict)