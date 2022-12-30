from openpyxl import Workbook
import openpyxl
from obj.record import Record
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook
import os

model1 = "glove"
model2 = "bert"

path1 = "error_predictions_" + model1 + ".xlsx"
path2 = "error_predictions_" + model2 + ".xlsx"

def getDbNameList():
    db_list = os.listdir('spider/database/')
    # for name in db_list:
    #     if name[0] != '.':
    #         db_list.remove(name)
    return db_list

def getSheetErrorList(wb1_sheet, wb2_sheet, sheet_list):
    for i in range(2, wb1_sheet.max_row + 1):
        q1 = wb1_sheet.cell(row=i, column=column_index_from_string('B')).value
        isCommon = False
        for j in range(2, wb2_sheet.max_row + 1):
            q2 = wb2_sheet.cell(row=j, column=column_index_from_string('B')).value
            if q1 == q2:
                isCommon = True
                break
        level = wb1_sheet.cell(row=i, column=column_index_from_string('A')).value
        gold = wb1_sheet.cell(row=i, column=column_index_from_string('C')).value
        predict = wb1_sheet.cell(row=i, column=column_index_from_string('D')).value
        record = [level, q1, gold, predict]

        if isCommon == False:
            sheet_list.append(record)

def outputResult(sheet_name, common_list, sheet1_list, sheet2_list):
    file_name = "compare_" + model1 + "_" + model2 + ".xlsx"
    wb = None
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()

    sheet = wb.create_sheet(sheet_name)

    sheet.cell(row=1, column=column_index_from_string('B')).value = 'Level'
    sheet.cell(row=1, column=column_index_from_string('C')).value = 'Question'
    sheet.cell(row=1, column=column_index_from_string('D')).value = 'Gold'
    sheet.cell(row=1, column=column_index_from_string('E')).value = 'Predict_' + model1
    sheet.cell(row=1, column=column_index_from_string('F')).value = 'Predict_' + model2

    row = 2
    sheet.cell(row=row, column=column_index_from_string('A')).value = 'Common Error'
    for item in common_list:
        sheet.cell(row=row, column=column_index_from_string('B')).value = item[0]
        sheet.cell(row=row, column=column_index_from_string('C')).value = item[1]
        sheet.cell(row=row, column=column_index_from_string('D')).value = item[2]
        sheet.cell(row=row, column=column_index_from_string('E')).value = item[3]
        sheet.cell(row=row, column=column_index_from_string('F')).value = item[4]
        row += 1

    row += 1
    sheet.cell(row=row, column=column_index_from_string('A')).value = model1 + " Error"
    for item in sheet1_list:
        sheet.cell(row=row, column=column_index_from_string('B')).value = item[0]
        sheet.cell(row=row, column=column_index_from_string('C')).value = item[1]
        sheet.cell(row=row, column=column_index_from_string('D')).value = item[2]
        sheet.cell(row=row, column=column_index_from_string('E')).value = item[3]
        row += 1

    row += 1
    sheet.cell(row=row, column=column_index_from_string('A')).value = model2 + " Error"

    for item in sheet2_list:
        sheet.cell(row=row, column=column_index_from_string('B')).value = item[0]
        sheet.cell(row=row, column=column_index_from_string('C')).value = item[1]
        sheet.cell(row=row, column=column_index_from_string('D')).value = item[2]
        sheet.cell(row=row, column=column_index_from_string('F')).value = item[3]
        row += 1

    wb.save(filename=file_name)

def getCommonError(dbList):
    wb1 = load_workbook(path1)
    wb2 = load_workbook(path2)

    # print(dbList)
    for name in dbList:
        try:
            wb1_sheet = wb1[name]
            wb2_sheet = wb2[name]
        except:
            continue

        common_list = []
        sheet1_list = []
        sheet2_list = []

        for i in range(2, wb1_sheet.max_row+1):
            q1 = wb1_sheet.cell(row=i, column=column_index_from_string('B')).value
            isCommon = False
            error_predict_q2 = None
            for j in range(2, wb2_sheet.max_row+1):
                q2 = wb2_sheet.cell(row=j, column=column_index_from_string('B')).value
                if q1 == q2:
                    isCommon = True
                    error_predict_q2 = wb2_sheet.cell(row=j, column=column_index_from_string('D')).value
                    break

            if isCommon:
                level = wb1_sheet.cell(row=i, column=column_index_from_string('A')).value
                gold = wb1_sheet.cell(row=i, column=column_index_from_string('C')).value
                predict = wb1_sheet.cell(row=i, column=column_index_from_string('D')).value
                record = [level, q1, gold, predict, error_predict_q2]
                common_list.append(record)
        getSheetErrorList(wb1_sheet, wb2_sheet, sheet1_list)
        getSheetErrorList(wb2_sheet, wb1_sheet, sheet2_list)

        if name == 'dog_kennels':
            print(sheet1_list)

        outputResult(name, common_list, sheet1_list, sheet2_list)

getCommonError(getDbNameList())

